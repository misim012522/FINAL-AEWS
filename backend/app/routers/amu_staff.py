"""AMU Staff endpoints: referrals (flagged enrollments), overview stats, reports."""
from datetime import datetime, timedelta, timezone
from bson import ObjectId
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pymongo.errors import ServerSelectionTimeoutError
import json
import csv
from io import StringIO
import secrets

from app.activity_log_utils import create_activity_log
from app.authz import get_current_actor
from app.database import get_db
from app.email_sender import send_student_support_email, send_needs_assessment_email
from app.notification_utils import create_notification
from app.schemas import NeedsAssessmentInvitationRequest, PublicNeedsAssessmentSubmission, ReferralEmailRequest
from app.ai_model import predict_student_risk
import os
import math
from typing import Dict, Any
try:
    import pickle
    import numpy as np
    import xgboost as xgb
except Exception:
    # model libs may not be installed in all environments; we'll fallback to heuristics
    xgb = None
    np = None
    pickle = None


def require_amu_staff_role(actor: dict = Depends(get_current_actor)):
    """Dependency to enforce AMU Staff access."""
    normalized_role = actor.get("role")
    if normalized_role != "amu-staff":
        raise HTTPException(status_code=403, detail="AMU Staff access required")
    return normalized_role


router = APIRouter(dependencies=[Depends(require_amu_staff_role)])
public_router = APIRouter()

REF_ID_SEP = "::"


def _normalize_referral_value(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _referral_identifier(doc: dict) -> str:
    student_email = _normalize_referral_value(doc.get("student_email")).lower()
    if student_email:
        return student_email
    return _normalize_referral_value(doc.get("student_id"))


def _referral_id(class_id: str, identifier: str) -> str:
    return f"{class_id}{REF_ID_SEP}{identifier}"


def _parse_ref_id(ref_id: str) -> tuple[str, str]:
    if REF_ID_SEP not in ref_id:
        raise HTTPException(status_code=400, detail="Invalid referral id")
    parts = ref_id.split(REF_ID_SEP, 1)
    return parts[0], parts[1]


def _find_referral_doc(db, class_id: str, identifier: str):
    ident = _normalize_referral_value(identifier)
    ident_lower = ident.lower()
    return db.enrollments.find_one({
        "class_id": class_id,
        "flagged_for_mentoring": True,
        "$or": [
            {"student_email": ident_lower},
            {"student_id": ident},
            {"student_id": ident_lower},
        ],
    })


def _serialize_support_routing(doc: dict) -> dict | None:
    verdict = doc.get("amu_final_verdict")
    if not isinstance(verdict, dict):
        return None

    saved_at = verdict.get("saved_at")
    if isinstance(saved_at, datetime):
        saved_at = saved_at.isoformat()

    return {
        "action": verdict.get("action"),
        "saved_by_id": verdict.get("saved_by_id"),
        "saved_by_name": verdict.get("saved_by_name"),
        "saved_at": saved_at,
    }


def _build_student_email(student_id: str | None) -> str | None:
    normalized = _normalize_referral_value(student_id)
    if not normalized:
        return None
    return f"{normalized}@student.buksu.edu.ph".lower()


def _get_referral_contact_email(doc: dict) -> str | None:
    student_email = _normalize_referral_value(doc.get("student_email")).lower()
    if student_email:
        return student_email
    return _build_student_email(doc.get("student_id"))


def _serialize_invitation(doc: dict) -> dict | None:
    invitation = doc.get("needs_assessment_invitation")
    if not isinstance(invitation, dict):
        return None

    def _format_dt(value):
        if isinstance(value, datetime):
            return value.isoformat()
        return value

    return {
        "status": invitation.get("status") or "not_sent",
        "token": invitation.get("token"),
        "email": invitation.get("email"),
        "sent_at": _format_dt(invitation.get("sent_at")),
        "submitted_at": _format_dt(invitation.get("submitted_at")),
        "last_error": invitation.get("last_error"),
    }


def _build_referral_reasons(doc: dict) -> list[str]:
    """Build referral reasons from instructor's selected checklist only."""
    reasons: list[str] = []
    
    # Check direct referral_reasons from instructor (dict submitted via form)
    referral_reasons_dict = doc.get("referral_reasons") or {}
    if isinstance(referral_reasons_dict, dict):
        support_map = {
            "on_probation_status": "On probation status",
            "grade_2_5_or_below": "Midterm grade is 2.50 or below",
            "gwa_2_5_or_below": "GWA is 2.5 or below",
            "low_midterm_performance": "Low midterm academic performance",
            "difficulty_catching_up": "Difficulty with catching up instructions",
            "financial_difficulties": "Financial difficulties",
            "physical_health_concerns": "Physical health concerns",
            "family_issues": "Family issues",
            "part_time_work_affecting_studies": "Part-time work affecting studies",
            "mental_health_concerns": "Mental health concerns",
        }
        for key, label in support_map.items():
            if referral_reasons_dict.get(key):
                reasons.append(label)
    
    # If no reasons from dict, check individual boolean fields as fallback
    if not reasons:
        individual_map = {
            "on_probation_status": "On probation status",
            "has_subject_grade_2_5": "Midterm grade is 2.50 or below",
            "gwa_2_5_or_below": "GWA is 2.5 or below",
            "low_midterm_academic_performance": "Low midterm academic performance",
            "difficulty_catching_up": "Difficulty with catching up instructions",
            "financial_difficulties": "Financial difficulties",
            "physical_health_concerns": "Physical health concerns",
            "family_issues": "Family issues",
            "part_time_work_affecting_studies": "Part-time work affecting studies",
            "mental_health_concerns": "Mental health concerns",
        }
        for key, label in individual_map.items():
            if doc.get(key):
                reasons.append(label)
    
    # Add instructor note if provided
    referral_note = (doc.get("referral_note") or "").strip()
    if referral_note:
        reasons.append(f"Instructor note: {referral_note}")

    # If no reasons were selected, show a default message
    if not reasons:
        reasons.append("Instructor referred the student for AMU support review.")

    return reasons


@router.get("/referrals")
def list_referrals(risk: str | None = None, search: str | None = None, actor: dict = Depends(get_current_actor)):
    """List all enrollments flagged for mentoring (referrals) with class and instructor info."""
    try:
        db = get_db()
        q = {"flagged_for_mentoring": True, "assigned_amu_staff_id": actor["id"]}
        if risk and risk in ("High", "Medium", "Low"):
            q["risk"] = risk
        cursor = db.enrollments.find(q).sort("referred_at", -1)
        search_lower = (search or "").strip().lower()
        out = []
        for doc in cursor:
            class_id = doc["class_id"]
            student_email = _normalize_referral_value(doc.get("student_email")).lower()
            student_id = _normalize_referral_value(doc.get("student_id"))
            identifier = _referral_identifier(doc)
            class_doc = db.classes.find_one({"_id": ObjectId(class_id)})
            if not class_doc:
                continue
            instructor_name = "Instructor"
            college = ""
            if class_doc.get("instructor_id"):
                inst_doc = db.instructor.find_one({"_id": ObjectId(class_doc["instructor_id"])})
                if inst_doc:
                    instructor_name = inst_doc.get("name") or instructor_name
                    college = (inst_doc.get("college") or "").strip()
            student_doc = db.students.find_one({"email": student_email}) if student_email else None
            student_name = (student_doc.get("name") if student_doc else None) or _normalize_referral_value(doc.get("student_name")) or student_id or student_email
            referred_at = doc.get("referred_at")
            referred_at_str = referred_at.strftime("%b %d, %Y") if referred_at else "—"
            if (
                search_lower
                and search_lower not in (student_email or "").lower()
                and search_lower not in (student_id or "").lower()
                and search_lower not in (student_name or "").lower()
            ):
                continue
            out.append({
                "id": _referral_id(class_id, identifier),
                "student_email": _get_referral_contact_email(doc),
                "student_id": student_id or None,
                "student_name": student_name,
                "class_id": class_id,
                "subject_code": class_doc.get("subject_code") or "",
                "subject_name": class_doc.get("subject_name") or "",
                "college": college,
                "risk": doc.get("risk") or "—",
                "referred_by": instructor_name,
                "referred_at": referred_at_str,
                "gpa": doc.get("gpa"),
                "attendance": doc.get("attendance"),
                "attendance_rate": doc.get("attendance_overall") if doc.get("attendance_overall") is not None else doc.get("attendance"),
                "midterm_grade": doc.get("midterm_grade"),
                "has_needs_assessment": bool(doc.get("needs_assessment")),
                "needs_assessment_uploaded_at": (
                    doc.get("needs_assessment_uploaded_at").strftime("%b %d, %Y")
                    if doc.get("needs_assessment_uploaded_at")
                    else None
                ),
                "needs_assessment_invitation": _serialize_invitation(doc),
                "risk_source": doc.get("risk_source"),
                "risk_source_label": doc.get("risk_source_label"),
                "risk_drivers": doc.get("risk_drivers") or [],
                "referral_note": doc.get("referral_note"),
                "assigned_amu_staff_id": doc.get("assigned_amu_staff_id"),
                "assigned_amu_staff_name": doc.get("assigned_amu_staff_name"),
                "assigned_amu_staff_college": doc.get("assigned_amu_staff_college"),
                "referral_reasons": _build_referral_reasons(doc),
                "amu_prediction": doc.get("amu_prediction"),
                "amu_prediction_generated_at": (
                    doc.get("amu_prediction_generated_at").isoformat()
                    if isinstance(doc.get("amu_prediction_generated_at"), datetime)
                    else doc.get("amu_prediction_generated_at")
                ),
                "support_routing": _serialize_support_routing(doc),
            })
        return out
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.delete("/referrals", status_code=200)
def delete_all_referrals(actor: dict = Depends(get_current_actor)):
    """Delete all referrals assigned to the current AMU staff member."""
    try:
        db = get_db()
        match_filter = {"flagged_for_mentoring": True, "assigned_amu_staff_id": actor["id"]}
        update_result = db.enrollments.update_many(
            match_filter,
            {
                "$set": {
                    "flagged_for_mentoring": False,
                },
                "$unset": {
                    "referral_note": "",
                    "referral_reasons": "",
                    "assigned_amu_staff_id": "",
                    "assigned_amu_staff_name": "",
                    "assigned_amu_staff_college": "",
                    "referred_at": "",
                    "needs_assessment": "",
                    "needs_assessment_uploaded_at": "",
                    "amu_prediction": "",
                    "amu_prediction_generated_at": "",
                    "support_routing": "",
                },
            },
        )
        create_activity_log(
            db,
            actor_id=actor["id"],
            actor_name=actor.get("name", "User"),
            role=actor["role"],
            action="delete_all_referrals",
            description=f"Deleted all referrals assigned to {actor.get('name', 'AMU staff')}.",
            target_type="referral",
            metadata={"deleted_count": update_result.modified_count},
        )
        return {"message": "All referrals deleted.", "deleted_count": update_result.modified_count}
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.get("/referrals/{ref_id:path}")
def get_referral(ref_id: str):
    """Get one referral by id (class_id::student_email or class_id::student_id)."""
    try:
        class_id, identifier = _parse_ref_id(ref_id)
        db = get_db()
        doc = _find_referral_doc(db, class_id, identifier)
        if not doc:
            raise HTTPException(status_code=404, detail="Referral not found.")
        class_doc = db.classes.find_one({"_id": ObjectId(class_id)})
        if not class_doc:
            raise HTTPException(status_code=404, detail="Class not found.")
        instructor_name = "Instructor"
        college = ""
        if class_doc.get("instructor_id"):
            inst_doc = db.instructor.find_one({"_id": ObjectId(class_doc["instructor_id"])})
            if inst_doc:
                instructor_name = inst_doc.get("name") or instructor_name
                college = (inst_doc.get("college") or "").strip()
        student_email = _normalize_referral_value(doc.get("student_email")).lower()
        student_id = _normalize_referral_value(doc.get("student_id"))
        student_doc = db.students.find_one({"email": student_email}) if student_email else None
        student_name = (student_doc.get("name") if student_doc else None) or _normalize_referral_value(doc.get("student_name")) or student_id or student_email
        referred_at = doc.get("referred_at")
        referred_at_str = referred_at.strftime("%b %d, %Y") if referred_at else "—"
        return {
            "id": ref_id,
            "student_email": _get_referral_contact_email(doc),
            "student_id": student_id or None,
            "student_name": student_name,
            "class_id": class_id,
            "subject_code": class_doc.get("subject_code") or "",
            "subject_name": class_doc.get("subject_name") or "",
            "college": college,
            "course": class_doc.get("subject_code") or "",
            "risk": doc.get("risk") or "—",
            "referred_by": instructor_name,
            "referred_at": referred_at_str,
            "gpa": doc.get("gpa"),
            "attendance": doc.get("attendance"),
            "midterm_grade": doc.get("midterm_grade"),
            "risk_probability_percent": doc.get("risk_probability_percent"),
            "previous_gpa": doc.get("previous_gpa"),
            "failed_subject_count": doc.get("failed_subject_count"),
            "has_needs_assessment": bool(doc.get("needs_assessment")),
            "needs_assessment": _normalize_needs_assessment(doc),
            "needs_assessment_invitation": _serialize_invitation(doc),
            "risk_source": doc.get("risk_source"),
            "risk_source_label": doc.get("risk_source_label"),
            "risk_drivers": doc.get("risk_drivers") or [],
            "academic_risk_drivers": doc.get("academic_risk_drivers") or [],
            "external_risk_drivers": doc.get("external_risk_drivers") or [],
            "referral_note": doc.get("referral_note"),
            "assigned_amu_staff_id": doc.get("assigned_amu_staff_id"),
            "assigned_amu_staff_name": doc.get("assigned_amu_staff_name"),
            "assigned_amu_staff_college": doc.get("assigned_amu_staff_college"),
            "referral_reasons": _build_referral_reasons(doc),
            "amu_prediction": doc.get("amu_prediction"),
            "amu_prediction_generated_at": (
                doc.get("amu_prediction_generated_at").isoformat()
                if isinstance(doc.get("amu_prediction_generated_at"), datetime)
                else doc.get("amu_prediction_generated_at")
            ),
            "support_routing": _serialize_support_routing(doc),
        }
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.post("/referrals/{ref_id:path}/notify")
def notify_referred_student(ref_id: str, body: ReferralEmailRequest, actor: dict = Depends(get_current_actor)):
    """Send a support email to a referred student."""
    try:
        class_id, identifier = _parse_ref_id(ref_id)
        db = get_db()
        doc = _find_referral_doc(db, class_id, identifier)
        if not doc:
            raise HTTPException(status_code=404, detail="Referral not found.")
        class_doc = db.classes.find_one({"_id": ObjectId(class_id)}) if ObjectId.is_valid(class_id) else None

        student_email = _normalize_referral_value(doc.get("student_email")).lower()
        if not student_email:
            raise HTTPException(status_code=400, detail="This referred student has no email address on file.")
        student_doc = db.students.find_one({"email": student_email})
        student_name = (student_doc.get("name") if student_doc else None) or _normalize_referral_value(doc.get("student_name")) or student_email
        sent, err = send_student_support_email(student_email, student_name, body.subject.strip(), body.message.strip())
        if not sent:
            raise HTTPException(status_code=500, detail=err or "Failed to send email.")

        create_notification(
            db,
            role="amu-staff",
            recipient_user_id=str(actor["id"]),
            title="Student support email sent",
            body=f"Support email sent to {student_email}.",
            type="report",
        )
        instructor_id = str(class_doc.get("instructor_id")) if class_doc and class_doc.get("instructor_id") else None
        create_notification(
            db,
            role="instructor",
            recipient_user_id=instructor_id,
            title="AMU reached out to a referred student",
            body=f"AMU staff sent a support email to {student_email}.",
            type="case",
        )
        create_activity_log(
            db,
            actor_id=actor["id"],
            actor_name=actor.get("name", "User"),
            role=actor["role"],
            action="notify_referred_student",
            description=f"Sent a support email to {student_email}.",
            target_type="referral",
            target_id=ref_id,
        )
        return {"message": f"Support email sent to {student_email}."}
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.post("/referrals/{ref_id:path}/needs-assessment")
async def upload_needs_assessment(ref_id: str, file: UploadFile = File(...), actor: dict = Depends(get_current_actor)):
    """Upload needs assessment file (CSV, Excel, or JSON) for a referred student."""
    try:
        class_id, identifier = _parse_ref_id(ref_id)
        db = get_db()
        doc = _find_referral_doc(db, class_id, identifier)
        if not doc:
            raise HTTPException(status_code=404, detail="Referral not found.")

        # Read file content
        content = await file.read()
        filename = file.filename or ""
        
        # Parse based on file type
        needs_assessment_data = {}
        if filename.endswith('.json'):
            try:
                needs_assessment_data = json.loads(content.decode('utf-8'))
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="Invalid JSON file.")
        elif filename.endswith('.csv'):
            try:
                text = content.decode('utf-8')
                reader = list(csv.reader(StringIO(text)))
                if len(reader) >= 3:
                    needs_assessment_data = _build_needs_assessment_payload_from_grouped_rows(
                        reader[0],
                        reader[1],
                        reader[2],
                    )
                elif len(reader) >= 2:
                    needs_assessment_data = {
                        _NEEDS_ASSESSMENT_FIELD_ALIASES.get(_slugify_needs_assessment_label(k), _slugify_needs_assessment_label(k)): _normalize_uploaded_value(v)
                        for k, v in zip(reader[0], reader[1])
                        if k not in (None, '')
                    }
                else:
                    raise HTTPException(status_code=400, detail="CSV file does not contain enough rows.")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid CSV file: {str(e)}")
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(content))
                ws = wb.active
                if ws.max_row >= 3:
                    section_row = [cell.value for cell in ws[1]]
                    field_row = [cell.value for cell in ws[2]]
                    data_row = [cell.value for cell in ws[3]]
                    needs_assessment_data = _build_needs_assessment_payload_from_grouped_rows(
                        section_row,
                        field_row,
                        data_row,
                    )
                elif ws.max_row >= 2:
                    headers = [cell.value for cell in ws[1]]
                    data_row = [cell.value for cell in ws[2]]
                    needs_assessment_data = {
                        _NEEDS_ASSESSMENT_FIELD_ALIASES.get(_slugify_needs_assessment_label(header), _slugify_needs_assessment_label(header)): _normalize_uploaded_value(value)
                        for header, value in zip(headers, data_row)
                        if header not in (None, '')
                    }
                else:
                    raise HTTPException(status_code=400, detail="Excel file does not contain enough rows.")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, Excel, or JSON.")

        normalized_needs_assessment = _normalize_needs_assessment_payload(needs_assessment_data)

        # Store needs assessment data in enrollment record
        db.enrollments.update_one(
            {"_id": doc["_id"]},
            {
                "$set": {
                    "needs_assessment": normalized_needs_assessment,
                    "needs_assessment_uploaded_at": datetime.now(timezone.utc),
                    "academic_challenge_score": normalized_needs_assessment.get("academic_challenge_score"),
                    "external_factor_score": normalized_needs_assessment.get("external_factor_score"),
                }
            },
        )
        create_activity_log(
            db,
            actor_id=actor["id"],
            actor_name=actor.get("name", "User"),
            role=actor["role"],
            action="upload_needs_assessment",
            description=f"Uploaded a needs assessment for referral {ref_id}.",
            target_type="referral",
            target_id=ref_id,
            metadata={"filename": filename, "field_count": len(normalized_needs_assessment)},
        )

        return {
            "message": "Needs assessment uploaded successfully.",
            "data_received": len(normalized_needs_assessment),
            "filename": filename,
        }
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


def _extract_features_from_enrollment(doc: dict) -> Dict[str, Any]:
    """Create a flat feature dict for prediction from enrollment document."""
    features: Dict[str, Any] = {}
    # Numeric metrics (prefer values stored on enrollment by instructor)
    # midterm_grade and attendance must come from enrollment as requested
    try:
        features['midterm_grade'] = float(doc.get('midterm_grade'))
    except Exception:
        features['midterm_grade'] = math.nan
    try:
        # attendance might be stored as percentage (number) or string
        att = doc.get('attendance')
        if isinstance(att, str) and att.endswith('%'):
            att = att.rstrip('%')
        features['attendance'] = float(att) if att not in (None, '') else math.nan
    except Exception:
        features['attendance'] = math.nan

    na = doc.get('needs_assessment') or {}
    # normalize keys to lowercase strings for lookups
    na_norm = {}
    if isinstance(na, dict):
        for k, v in na.items():
            if k is None:
                continue
            na_norm[str(k).strip().lower()] = v

    try:
        previous_gpa_value = na_norm.get('previous_gpa')
        if previous_gpa_value in (None, ''):
            previous_gpa_value = doc.get('previous_gpa')
        features['previous_gpa'] = float(previous_gpa_value) if previous_gpa_value not in (None, '') else math.nan
    except Exception:
        features['previous_gpa'] = math.nan
    try:
        failed_subject_count_value = na_norm.get('failed_subject_count')
        if failed_subject_count_value in (None, ''):
            failed_subject_count_value = doc.get('failed_subject_count')
        features['failed_subject_count'] = int(float(failed_subject_count_value or 0))
    except Exception:
        features['failed_subject_count'] = 0

    def parse_num_from_na(*keys):
        for key in keys:
            v = na_norm.get(key)
            if v is None:
                continue
            try:
                return float(v)
            except Exception:
                # try extract digits
                try:
                    s = str(v)
                    s2 = ''.join(ch for ch in s if (ch.isdigit() or ch == '.' or ch == '-'))
                    return float(s2) if s2 else math.nan
                except Exception:
                    continue
        return math.nan

    # Common numeric fields
    features['self_reported_difficulty'] = parse_num_from_na('difficulty', 'self_reported_difficulty', 'difficulty_level')
    features['hours_studied_per_week'] = parse_num_from_na('hours_studied', 'hours_per_week', 'study_hours')

    # Binary / categorical flags from needs assessment
    def parse_bool_na(*keys):
        for key in keys:
            v = na_norm.get(key)
            if v is None:
                continue
            if isinstance(v, bool):
                return 1 if v else 0
            s = str(v).strip().lower()
            if s in ('yes', 'y', 'true', '1'):
                return 1
            if s in ('no', 'n', 'false', '0'):
                return 0
        return 0

    features['has_personal_issues'] = parse_bool_na('personal_issues', 'personal_problems', 'personal_issues_reported')
    features['has_financial_issues'] = parse_bool_na('financial_issues', 'financial_difficulties', 'financial_concerns')
    features['has_health_issues'] = parse_bool_na('health_issues', 'medical_issues')
    features['has_internet_issues'] = parse_bool_na('internet_issues', 'connectivity_issues')

    # Attempt to extract any free-text notes length as a weak signal
    note = na_norm.get('notes') or na_norm.get('comments') or na_norm.get('description') or ''
    try:
        features['notes_length'] = float(len(str(note)))
    except Exception:
        features['notes_length'] = 0.0

    return features


def _normalize_needs_assessment(doc: dict) -> dict[str, Any]:
    payload = doc.get("needs_assessment") or {}
    normalized: dict[str, Any] = {}
    if isinstance(payload, dict):
        for key, value in payload.items():
            if key is None:
                continue
            normalized[str(key).strip().lower()] = value
    return normalized


def _serialize_public_invitation(doc: dict) -> dict[str, Any]:
    invitation = doc.get("needs_assessment_invitation") or {}
    student_name = _normalize_referral_value(doc.get("student_name")) or _normalize_referral_value(doc.get("student_id")) or "Student"
    return {
        "student_name": student_name,
        "student_id": _normalize_referral_value(doc.get("student_id")) or None,
        "student_email": _get_referral_contact_email(doc),
        "status": invitation.get("status") or ("completed" if doc.get("needs_assessment") else "sent"),
        "submitted_at": invitation.get("submitted_at").isoformat() if isinstance(invitation.get("submitted_at"), datetime) else invitation.get("submitted_at"),
        "can_submit": not bool(doc.get("needs_assessment")),
    }


def _is_missing_required_needs_assessment_value(field: dict[str, Any], value: Any) -> bool:
    if field.get("type") == "boolean":
        return value is not True
    if field.get("type") == "number":
        return value in {"", None}
    return str(value or "").strip() == ""


def _collect_missing_required_needs_assessment_fields(payload: dict[str, Any], form: dict[str, Any] | None) -> list[str]:
    missing: list[str] = []
    if not isinstance(form, dict):
        return missing
    for section in form.get("sections") or []:
        if not isinstance(section, dict):
            continue
        for field in section.get("fields") or []:
            if not isinstance(field, dict) or not field.get("required") or field.get("active") is False:
                continue
            field_name = str(field.get("name") or "").strip()
            if not field_name:
                continue
            if _is_missing_required_needs_assessment_value(field, payload.get(field_name)):
                missing.append(str(field.get("label") or field_name))
    return missing


def _slugify_needs_assessment_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    cleaned = []
    previous_underscore = False
    for char in text:
        if char.isalnum():
            cleaned.append(char)
            previous_underscore = False
        else:
            if not previous_underscore:
                cleaned.append("_")
                previous_underscore = True
    return "".join(cleaned).strip("_")


_NEEDS_ASSESSMENT_FIELD_ALIASES = {
    "student_id": "student_id",
    "name": "name",
    "course_year": "course_year",
    "college_department": "college_department",
    "campus": "campus",
    "email_address": "email_address",
    "admission_type": "admission_type",
    "academic_adviser": "academic_adviser",
    "on_probationary_status": "on_probationary_status",
    "at_least_one_subject_has_a_grade_of_2_5": "grade_2_5_or_below",
    "gwa_is_2_5_lower_or_below": "gwa_2_5_or_below",
    "low_midterm_academic_performance": "low_midterm_academic_performance",
    "difficulty_with_catching_up_instructions": "difficulty_catching_up",
    "previous_year_semester": "previous_year_semester",
    "previous_gpa": "previous_gpa",
    "no_of_subjects_failed_if_any": "failed_subject_count",
    "regular_attendance": "regular_attendance",
    "frequently_absent_late": "frequently_absent_or_late",
    "tutoring_sessions": "tutoring_sessions",
    "peer_mentoring": "peer_mentoring",
    "faculty_consulation": "faculty_consultation",
    "faculty_consultation": "faculty_consultation",
    "counselling_sessions": "counselling_sessions",
    "none": "no_previous_support",
    "difficulty_in_understanding_lectures": "difficulty_understanding_lectures",
    "struggles_with_specific_subjects": "struggles_specific_subjects",
    "weak_study_habits_or_time_management": "weak_study_habits_time_management",
    "low_motivation_or_engagement": "low_motivation_engagement",
    "poor_comprehension_or_writing_skills": "poor_comprehension_writing_skills",
    "financial_difficulties": "financial_difficulties",
    "physical_health_related_concerns": "physical_health_concerns",
    "family_issues": "family_issues",
    "part_time_work_affecting_studies": "part_time_work_affecting_studies",
    "mental_health_related_concerns": "mental_health_concerns",
    "internet_connectivity_issues": "internet_issues",
    "internet_issues": "internet_issues",
    "connectivity_issues": "internet_issues",
}


def _normalize_uploaded_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return ""
        lowered = stripped.lower()
        if lowered in {"yes", "y", "true"}:
            return 1
        if lowered in {"no", "n", "false"}:
            return 0
        return stripped
    return value


def _normalize_needs_assessment_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    grouped_sections = payload.get("_grouped_sections") if isinstance(payload.get("_grouped_sections"), dict) else {}
    column_map = payload.get("_column_map") if isinstance(payload.get("_column_map"), list) else []

    for raw_key, raw_value in payload.items():
        if raw_key in {"_grouped_sections", "_column_map"} or raw_key is None:
            continue
        field_key = _slugify_needs_assessment_label(raw_key)
        normalized_key = _NEEDS_ASSESSMENT_FIELD_ALIASES.get(field_key, field_key)
        normalized[normalized_key] = _normalize_uploaded_value(raw_value)

    academic_fields = [
        "difficulty_understanding_lectures",
        "struggles_specific_subjects",
        "weak_study_habits_time_management",
        "low_motivation_engagement",
        "poor_comprehension_writing_skills",
    ]
    external_fields = [
        "financial_difficulties",
        "physical_health_concerns",
        "family_issues",
        "part_time_work_affecting_studies",
        "mental_health_concerns",
        "internet_issues",
    ]

    normalized["academic_challenge_score"] = sum(1 for field in academic_fields if _is_truthy_value(normalized.get(field)))
    normalized["external_factor_score"] = sum(1 for field in external_fields if _is_truthy_value(normalized.get(field)))

    compatibility_aliases = {
        "on_probation_status": normalized.get("on_probationary_status"),
        "has_subject_grade_2_5": normalized.get("grade_2_5_or_below"),
        "difficulty_catching_up_instructions": normalized.get("difficulty_catching_up"),
        "low_midterm_performance": normalized.get("low_midterm_academic_performance"),
        "health_issues": normalized.get("physical_health_concerns"),
    }
    for alias_key, alias_value in compatibility_aliases.items():
        if alias_value is not None:
            normalized[alias_key] = alias_value

    if grouped_sections:
        normalized["_grouped_sections"] = grouped_sections
    if column_map:
        normalized["_column_map"] = column_map
    return normalized


def _build_needs_assessment_payload_from_grouped_rows(
    section_row: list[Any],
    field_row: list[Any],
    data_row: list[Any],
) -> dict[str, Any]:
    current_section = ""
    flat_data: dict[str, Any] = {}
    grouped_data: dict[str, dict[str, Any]] = {}
    column_map: list[dict[str, str]] = []

    for raw_section, raw_field, raw_value in zip(section_row, field_row, data_row):
        if raw_section not in (None, ""):
            current_section = str(raw_section).strip()
        if raw_field in (None, ""):
            continue

        field_key = _slugify_needs_assessment_label(raw_field)
        normalized_key = _NEEDS_ASSESSMENT_FIELD_ALIASES.get(field_key, field_key)
        normalized_value = _normalize_uploaded_value(raw_value)
        section_key = _slugify_needs_assessment_label(current_section) or "ungrouped"

        flat_data[normalized_key] = normalized_value
        grouped_data.setdefault(section_key, {})[normalized_key] = normalized_value
        column_map.append(
            {
                "section": current_section or "Ungrouped",
                "section_key": section_key,
                "label": str(raw_field).strip(),
                "field_key": normalized_key,
            }
        )

    payload = {
        **flat_data,
        "_grouped_sections": grouped_data,
        "_column_map": column_map,
    }
    return _normalize_needs_assessment_payload(payload)


def _is_truthy_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value or "").strip().lower()
    return text in {"yes", "y", "true", "1", "present", "checked"}


def _build_prediction_explanations(doc: dict, result: dict[str, Any]) -> list[str]:
    explanations: list[str] = []
    features = result.get("features") or {}
    needs_assessment = _normalize_needs_assessment(doc)

    midterm_grade = features.get("midterm_grade")
    if isinstance(midterm_grade, (int, float)):
        if midterm_grade >= 3:
            explanations.append(f"Midterm grade is {midterm_grade:.2f}, which is in the low-performing range.")
        elif midterm_grade > 0 and midterm_grade <= 75:
            explanations.append(f"Midterm grade is only {midterm_grade:.2f}.")

    attendance_rate = features.get("attendance_rate")
    if isinstance(attendance_rate, (int, float)) and attendance_rate < 85:
        explanations.append(f"Attendance is low at {attendance_rate:.0f}%.")

    failed_subject_count = features.get("failed_subject_count")
    if isinstance(failed_subject_count, (int, float)) and int(failed_subject_count) > 0:
        count = int(failed_subject_count)
        explanations.append(f"The student has {count} subject{'s' if count != 1 else ''} with failing or concerning marks.")

    previous_gpa = features.get("previous_gpa")
    if isinstance(previous_gpa, (int, float)) and previous_gpa >= 2.25:
        explanations.append(f"Previous GPA is concerning at {previous_gpa:.2f}.")

    external_flag_map = [
        (("part_time_work_affecting_studies", "working_student", "working student"), "The needs assessment indicates the student is a working student."),
        (("financial_difficulties", "financial_issues", "financial difficulties"), "The needs assessment reports financial difficulties."),
        (("family_issues", "family concerns", "family_issue"), "The needs assessment reports family-related concerns."),
        (("mental_health_concerns", "mental health", "mental_health"), "The needs assessment reports mental health concerns."),
        (("physical_health_concerns", "health_issues", "medical_issues"), "The needs assessment reports health concerns."),
        (("internet_issues", "connectivity_issues", "internet problems"), "The needs assessment reports internet or connectivity problems."),
    ]
    for keys, message in external_flag_map:
        if any(_is_truthy_value(needs_assessment.get(key)) for key in keys):
            explanations.append(message)

    for detail in result.get("risk_drivers") or []:
        if detail and detail not in explanations:
            explanations.append(str(detail))

    for signal in result.get("top_contributing_signals") or []:
        detail = (signal or {}).get("detail")
        if detail and detail not in explanations:
            explanations.append(str(detail))

    seen: set[str] = set()
    deduped: list[str] = []
    for item in explanations:
        normalized = item.strip()
        if not normalized:
            continue
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(normalized)
    return deduped


def _build_weight_explanations(doc: dict, result: dict[str, Any]) -> tuple[list[str], list[str]]:
    features = result.get("features") or {}
    needs_assessment = _normalize_needs_assessment(doc)

    academic_items: list[str] = []
    external_items: list[str] = []

    midterm_grade = features.get("midterm_grade")
    if isinstance(midterm_grade, (int, float)) and midterm_grade >= 2.0:
        academic_items.append(f"Midterm grade is {midterm_grade:.2f}, which signals academic difficulty.")

    attendance_rate = features.get("attendance_rate")
    if isinstance(attendance_rate, (int, float)) and attendance_rate < 85:
        academic_items.append(f"Attendance is only {attendance_rate:.0f}%, which points to academic risk.")

    failed_subject_count = features.get("failed_subject_count")
    if isinstance(failed_subject_count, (int, float)) and int(failed_subject_count) > 0:
        count = int(failed_subject_count)
        academic_items.append(f"The student has {count} subject{'s' if count != 1 else ''} with failing or concerning marks.")

    previous_gpa = features.get("previous_gpa")
    if isinstance(previous_gpa, (int, float)) and previous_gpa >= 2.25:
        academic_items.append(f"Previous GPA is {previous_gpa:.2f}, which adds to the academic concern.")

    academic_flag_map = [
        (("difficulty_understanding_lectures",), "The needs assessment reports difficulty understanding lectures."),
        (("struggles_specific_subjects",), "The needs assessment reports struggles with specific subjects."),
        (("weak_study_habits_time_management",), "The needs assessment reports weak study habits or time management."),
        (("low_motivation_engagement",), "The needs assessment reports low motivation or engagement."),
        (("poor_comprehension_writing_skills",), "The needs assessment reports poor comprehension or writing skills."),
    ]
    for keys, message in academic_flag_map:
        if any(_is_truthy_value(needs_assessment.get(key)) for key in keys):
            academic_items.append(message)

    external_flag_map = [
        (("part_time_work_affecting_studies", "working_student"), "The needs assessment indicates the student is a working student."),
        (("financial_difficulties",), "The needs assessment reports financial difficulties."),
        (("physical_health_concerns", "health_issues", "medical_issues"), "The needs assessment reports health concerns."),
        (("family_issues",), "The needs assessment reports family-related concerns."),
        (("mental_health_concerns",), "The needs assessment reports mental health concerns."),
        (("internet_issues", "connectivity_issues"), "The needs assessment reports internet or connectivity problems."),
    ]
    for keys, message in external_flag_map:
        if any(_is_truthy_value(needs_assessment.get(key)) for key in keys):
            external_items.append(message)

    def _dedupe(items: list[str]) -> list[str]:
        seen: set[str] = set()
        result_items: list[str] = []
        for item in items:
            key = item.strip().lower()
            if not key or key in seen:
                continue
            seen.add(key)
            result_items.append(item.strip())
        return result_items

    return _dedupe(academic_items), _dedupe(external_items)


def _load_xgb_model(path: str):
    if not xgb or not pickle:
        return None
    if not os.path.exists(path):
        # If pickle not present, try to find common JSON model files and create pickle
        candidates = [
            os.path.join(os.path.dirname(__file__), '..', '..', 'xgboost_student_risk.json'),
            os.path.join(os.path.dirname(__file__), '..', '..', 'xgboost_student_risk_midterm_endterm.json'),
            os.path.join(os.path.dirname(__file__), '..', '..', 'xgboost_student_risk_early_warning.json'),
        ]
        for c in candidates:
            if os.path.exists(c):
                try:
                    booster = xgb.Booster()
                    booster.load_model(c)
                    # ensure models dir
                    outdir = os.path.dirname(path)
                    os.makedirs(outdir, exist_ok=True)
                    with open(path, 'wb') as f:
                        pickle.dump(booster, f)
                    return booster
                except Exception:
                    continue
        return None
    try:
        with open(path, 'rb') as f:
            model = pickle.load(f)
        return model
    except Exception:
        return None


def _predict_with_model(model, features: Dict[str, Any]):
    # expects model supports .predict_proba or xgboost Booster
    try:
        if hasattr(model, 'predict_proba') and np is not None:
            arr = np.array([list(features.values())], dtype=float)
            probs = model.predict_proba(arr)[0]
            # assume binary: probs[1] is risk prob
            prob = float(probs[1]) if len(probs) > 1 else float(probs[0])
            return prob
        elif xgb and isinstance(model, xgb.Booster):
            arr = np.array([list(features.values())], dtype=float)
            dmat = xgb.DMatrix(arr)
            pred = model.predict(dmat)
            return float(pred[0])
    except Exception:
        return None


@router.post("/referrals/{ref_id:path}/predict")
def predict_referral(ref_id: str, actor: dict = Depends(get_current_actor)):
    """Generate a risk prediction for a referred student using needs assessment and enrollment metrics.

    Returns the AMU outcome label, explanation factors, and probability metadata.
    """
    try:
        class_id, identifier = _parse_ref_id(ref_id)
        db = get_db()
        doc = _find_referral_doc(db, class_id, identifier)
        if not doc:
            raise HTTPException(status_code=404, detail="Referral not found.")
        if not doc.get("needs_assessment"):
            return {
                "prediction_status": "awaiting_needs_assessment",
                "probability": None,
                "prediction_label": "Awaiting needs assessment",
                "prediction_basis": "awaiting_needs_assessment",
                "message": "Upload a needs assessment first before AMU can run the prediction.",
                "contributing_factors": [],
                "academic_weight_reasons": [],
                "external_weight_reasons": [],
                "academic_weight_score": 0,
                "external_weight_score": 0,
                "risk_source": None,
                "risk_source_label": None,
                "used_model": False,
                "model_source": None,
                "support_routing": _serialize_support_routing(doc),
            }

        normalized_needs_assessment = _normalize_needs_assessment_payload(doc.get("needs_assessment") or {})
        doc["needs_assessment"] = normalized_needs_assessment
        doc["academic_challenge_score"] = normalized_needs_assessment.get("academic_challenge_score")
        doc["external_factor_score"] = normalized_needs_assessment.get("external_factor_score")

        result = predict_student_risk(doc)
        probability = result.get("probability")
        if probability is None:
            probability = (result.get("probability_percent") or 0) / 100
        factors = _build_prediction_explanations(doc, result)
        academic_weight_reasons, external_weight_reasons = _build_weight_explanations(doc, result)
        risk_source = result.get("risk_source")
        academic_weight_score = result.get("academic_weight_score") or 0
        external_weight_score = result.get("external_weight_score") or 0
        prediction_label = "External Factor" if risk_source == "external_factors" else "Academic Problem"

        # Save prediction to enrollment doc
        db.enrollments.update_one(
            {"_id": doc["_id"]},
            {"$set": {
                "needs_assessment": normalized_needs_assessment,
                "academic_challenge_score": doc.get("academic_challenge_score"),
                "external_factor_score": doc.get("external_factor_score"),
                "amu_prediction": {
                    "prediction_status": "ready",
                    "probability": probability,
                    "prediction_label": prediction_label,
                    "prediction_basis": "weight_score_comparison",
                    "factors": factors,
                    "academic_weight_reasons": academic_weight_reasons,
                    "external_weight_reasons": external_weight_reasons,
                    "academic_weight_score": academic_weight_score,
                    "external_weight_score": external_weight_score,
                    "risk_source": result.get("risk_source"),
                    "risk_source_label": result.get("risk_source_label"),
                    "risk_drivers": result.get("risk_drivers") or [],
                    "academic_risk_drivers": result.get("academic_risk_drivers") or [],
                    "external_risk_drivers": result.get("external_risk_drivers") or [],
                    "top_contributing_signals": result.get("top_contributing_signals") or [],
                    "features": result.get("features") or {},
                    "used_model": result.get("model_path") is not None,
                    "model_source": result.get("model_source"),
                },
                "amu_prediction_generated_at": datetime.now(timezone.utc),
            }}
        )
        create_activity_log(
            db,
            actor_id=actor["id"],
            actor_name=actor.get("name", "User"),
            role=actor["role"],
            action="predict_referral",
            description=f"Generated AMU prediction for referral {ref_id}.",
            target_type="referral",
            target_id=ref_id,
            metadata={"prediction_label": prediction_label, "risk_source": result.get("risk_source")},
        )

        # Return whether we used the model or heuristic for easier debugging
        return {
            "prediction_status": "ready",
            "probability": probability,
            "prediction_label": prediction_label,
            "prediction_basis": "weight_score_comparison",
            "contributing_factors": factors,
            "academic_weight_reasons": academic_weight_reasons,
            "external_weight_reasons": external_weight_reasons,
            "academic_weight_score": academic_weight_score,
            "external_weight_score": external_weight_score,
            "risk_source": result.get("risk_source"),
            "risk_source_label": result.get("risk_source_label"),
            "risk_drivers": result.get("risk_drivers") or [],
            "academic_risk_drivers": result.get("academic_risk_drivers") or [],
            "external_risk_drivers": result.get("external_risk_drivers") or [],
            "top_contributing_signals": result.get("top_contributing_signals") or [],
            "features": result.get("features") or {},
            "used_model": result.get("model_path") is not None,
            "model_source": result.get("model_source"),
            "support_routing": _serialize_support_routing(doc),
        }
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/referrals/{ref_id:path}/support-routing")
def save_support_routing(ref_id: str, payload: dict, actor: dict = Depends(get_current_actor)):
    """Save AMU staff's support routing after prediction review."""
    try:
        class_id, identifier = _parse_ref_id(ref_id)
        db = get_db()
        doc = _find_referral_doc(db, class_id, identifier)
        if not doc:
            raise HTTPException(status_code=404, detail="Referral not found.")

        action = str((payload or {}).get("action") or "").strip()
        allowed_actions = {
            "mentoring",
            "counselling",
            "both_mentoring_and_counselling",
            "monitoring_only",
            "other_support",
        }
        if action not in allowed_actions:
            raise HTTPException(status_code=400, detail="Choose a valid support routing action.")

        verdict = {
            "action": action,
            "saved_by_id": actor.get("id"),
            "saved_by_name": actor.get("name"),
            "saved_at": datetime.now(timezone.utc),
        }

        db.enrollments.update_one(
            {"_id": doc["_id"]},
            {"$set": {"amu_final_verdict": verdict}},
        )
        create_activity_log(
            db,
            actor_id=actor["id"],
            actor_name=actor.get("name", "User"),
            role=actor["role"],
            action="save_support_routing",
            description=f"Saved support routing '{action}' for referral {ref_id}.",
            target_type="referral",
            target_id=ref_id,
            metadata={"action": action},
        )

        return {
            "message": "Support routing saved successfully.",
            "support_routing": _serialize_support_routing({"amu_final_verdict": verdict}),
        }
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.get("/overview")
def get_overview(actor: dict = Depends(get_current_actor)):
    """Return counts for the AMU overview dashboard."""
    try:
        db = get_db()
        base_match = {"flagged_for_mentoring": True, "assigned_amu_staff_id": actor["id"]}
        referrals_count = db.enrollments.count_documents(base_match)
        courses_with_referrals = len(db.enrollments.distinct("class_id", base_match))
        needs_assessment_queue = db.enrollments.count_documents({
            **base_match,
            "amu_final_verdict": {"$exists": False},
        })
        prediction_ready = db.enrollments.count_documents({
            **base_match,
            "needs_assessment": {"$exists": True, "$ne": None},
            "amu_final_verdict": {"$exists": False},
        })
        return {
            "referrals_count": referrals_count,
            "courses_monitored": courses_with_referrals,
            "needs_assessment_queue": needs_assessment_queue,
            "prediction_ready": prediction_ready,
        }
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


def _send_needs_assessment_invitation_for_doc(db, doc: dict, custom_message: str | None):
    student_email = _get_referral_contact_email(doc)
    if not student_email:
        raise HTTPException(status_code=400, detail="This referred student has no student ID or email on file.")

    student_name = _normalize_referral_value(doc.get("student_name")) or _normalize_referral_value(doc.get("student_id")) or student_email
    existing_invitation = doc.get("needs_assessment_invitation") or {}
    token = existing_invitation.get("token") or secrets.token_urlsafe(32)
    frontend_url = os.getenv("FRONTEND_URL", "http://localhost:5173").rstrip("/")
    form_link = f"{frontend_url}/needs-assessment/{token}"
    now = datetime.now(timezone.utc)

    sent, err = send_needs_assessment_email(
        student_email,
        student_name,
        form_link,
        custom_message,
    )
    if not sent:
        db.enrollments.update_one(
            {"_id": doc["_id"]},
            {
                "$set": {
                    "needs_assessment_invitation": {
                        **existing_invitation,
                        "token": token,
                        "email": student_email,
                        "status": "failed",
                        "last_error": err or "Failed to send email.",
                    }
                }
            },
        )
        raise HTTPException(status_code=500, detail=err or "Failed to send email.")

    invitation = {
        "token": token,
        "email": student_email,
        "status": "completed" if doc.get("needs_assessment") else "sent",
        "sent_at": now,
        "submitted_at": existing_invitation.get("submitted_at"),
        "last_error": None,
    }
    db.enrollments.update_one({"_id": doc["_id"]}, {"$set": {"needs_assessment_invitation": invitation}})
    return {
        "student_email": student_email,
        "invitation": invitation,
        "form_link": form_link,
    }


@router.post("/referrals/{ref_id:path}/needs-assessment/send")
def send_needs_assessment_invitation(
    ref_id: str,
    body: NeedsAssessmentInvitationRequest,
    actor: dict = Depends(get_current_actor),
):
    """Send or resend a tokenized needs assessment form link to a referred student."""
    try:
        class_id, identifier = _parse_ref_id(ref_id)
        db = get_db()
        doc = _find_referral_doc(db, class_id, identifier)
        if not doc:
            raise HTTPException(status_code=404, detail="Referral not found.")

        result = _send_needs_assessment_invitation_for_doc(db, doc, body.custom_message)

        create_activity_log(
            db,
            actor_id=actor["id"],
            actor_name=actor.get("name", "User"),
            role=actor["role"],
            action="send_needs_assessment_invitation",
            description=f"Sent needs assessment form to {result['student_email']}.",
            target_type="referral",
            target_id=ref_id,
        )
        return {
            "message": f"Needs assessment form sent to {result['student_email']}.",
            "invitation": _serialize_invitation({"needs_assessment_invitation": result["invitation"]}),
            "form_link": result["form_link"],
        }
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.post("/needs-assessments/send-all")
def send_all_needs_assessment_invitations(
    body: NeedsAssessmentInvitationRequest,
    actor: dict = Depends(get_current_actor),
):
    """Send or resend tokenized needs assessment form links to all current referrals for this AMU staff member."""
    try:
        db = get_db()
        cursor = db.enrollments.find({
            "flagged_for_mentoring": True,
            "assigned_amu_staff_id": actor["id"],
        }).sort("referred_at", -1)

        sent_count = 0
        failed: list[dict] = []

        for doc in cursor:
            ref_id = _referral_id(doc["class_id"], _referral_identifier(doc))
            try:
                result = _send_needs_assessment_invitation_for_doc(db, doc, body.custom_message)
                sent_count += 1
                create_activity_log(
                    db,
                    actor_id=actor["id"],
                    actor_name=actor.get("name", "User"),
                    role=actor["role"],
                    action="send_needs_assessment_invitation",
                    description=f"Sent needs assessment form to {result['student_email']}.",
                    target_type="referral",
                    target_id=ref_id,
                )
            except HTTPException as exc:
                failed.append({
                    "ref_id": ref_id,
                    "student_name": _normalize_referral_value(doc.get("student_name")) or _normalize_referral_value(doc.get("student_id")) or _get_referral_contact_email(doc) or "Student",
                    "detail": exc.detail if isinstance(exc.detail, str) else "Failed to send email.",
                })

        summary = f"Needs assessment forms sent to {sent_count} referred student{'s' if sent_count != 1 else ''}."
        if failed:
            summary += f" {len(failed)} failed."

        return {
            "message": summary,
            "sent_count": sent_count,
            "failed_count": len(failed),
            "failed": failed,
        }
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.get("/needs-assessments/export")
def export_needs_assessments(actor: dict = Depends(get_current_actor)):
    """Export completed needs assessment responses for the current AMU staff member."""
    try:
        db = get_db()
        cursor = db.enrollments.find(
            {
                "flagged_for_mentoring": True,
                "assigned_amu_staff_id": actor["id"],
                "needs_assessment": {"$exists": True, "$ne": None},
            }
        ).sort("referred_at", -1)

        fieldnames = [
            "student_id",
            "student_email",
            "student_name",
            "subject_code",
            "subject_name",
            "referred_at",
            "submitted_at",
            "previous_gpa",
            "failed_subject_count",
            "regular_attendance",
            "frequently_absent_or_late",
            "tutoring_sessions",
            "peer_mentoring",
            "faculty_consultation",
            "counselling_sessions",
            "no_previous_support",
            "difficulty_understanding_lectures",
            "struggles_specific_subjects",
            "weak_study_habits_time_management",
            "low_motivation_engagement",
            "poor_comprehension_writing_skills",
            "financial_difficulties",
            "physical_health_concerns",
            "family_issues",
            "part_time_work_affecting_studies",
            "mental_health_concerns",
            "internet_issues",
            "notes",
        ]

        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fieldnames)
        writer.writeheader()
        for doc in cursor:
            needs = doc.get("needs_assessment") or {}
            invitation = doc.get("needs_assessment_invitation") or {}
            class_doc = db.classes.find_one({"_id": ObjectId(doc["class_id"])}) if ObjectId.is_valid(doc["class_id"]) else None
            student_email = _get_referral_contact_email(doc)
            referred_at = doc.get("referred_at")
            submitted_at = invitation.get("submitted_at") or doc.get("needs_assessment_uploaded_at")
            writer.writerow({
                "student_id": _normalize_referral_value(doc.get("student_id")),
                "student_email": student_email or "",
                "student_name": _normalize_referral_value(doc.get("student_name")),
                "subject_code": class_doc.get("subject_code") if class_doc else "",
                "subject_name": class_doc.get("subject_name") if class_doc else "",
                "referred_at": referred_at.isoformat() if isinstance(referred_at, datetime) else "",
                "submitted_at": submitted_at.isoformat() if isinstance(submitted_at, datetime) else (submitted_at or ""),
                **{key: needs.get(key, "") for key in fieldnames if key not in {
                    "student_id", "student_email", "student_name", "subject_code", "subject_name", "referred_at", "submitted_at"
                }},
            })

        filename = f"needs-assessment-responses-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@router.get("/reports")
def get_reports(actor: dict = Depends(get_current_actor)):
    """Monthly summary of referrals and saved AMU support routing outcomes."""
    try:
        db = get_db()
        base_match = {"flagged_for_mentoring": True, "assigned_amu_staff_id": actor["id"]}
        pipeline_refs = [
            {"$match": {**base_match, "referred_at": {"$exists": True, "$ne": None}}},
            {"$group": {"_id": {"year": {"$year": "$referred_at"}, "month": {"$month": "$referred_at"}}, "count": {"$sum": 1}}},
            {"$sort": {"_id.year": -1, "_id.month": -1}},
            {"$limit": 12},
        ]
        refs_by_month = list(db.enrollments.aggregate(pipeline_refs))
        out = []
        for r in refs_by_month:
            year, month = r["_id"]["year"], r["_id"]["month"]
            dt = datetime(year, month, 1, tzinfo=timezone.utc)
            period = dt.strftime("%B %Y")
            out.append({
                "period": period,
                "referrals": r["count"],
            })
        if not out:
            out.append({
                "period": datetime.now(timezone.utc).strftime("%B %Y"),
                "referrals": 0,
            })

        verdict_counts = {
            "mentoring": db.enrollments.count_documents({**base_match, "amu_final_verdict.action": "mentoring"}),
            "counselling": db.enrollments.count_documents({**base_match, "amu_final_verdict.action": "counselling"}),
            "both_mentoring_and_counselling": db.enrollments.count_documents({**base_match, "amu_final_verdict.action": "both_mentoring_and_counselling"}),
            "monitoring_only": db.enrollments.count_documents({**base_match, "amu_final_verdict.action": "monitoring_only"}),
            "other_support": db.enrollments.count_documents({**base_match, "amu_final_verdict.action": "other_support"}),
        }
        verdict_counts["total_with_routing"] = sum(verdict_counts.values())

        return {
            "history": out,
            "support_routing_summary": verdict_counts,
        }
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@public_router.get("/needs-assessments/{token}")
def get_public_needs_assessment(token: str):
    """Return invitation metadata for a student needs assessment form."""
    try:
        db = get_db()
        doc = db.enrollments.find_one({"needs_assessment_invitation.token": token, "flagged_for_mentoring": True})
        if not doc:
            raise HTTPException(status_code=404, detail="Needs assessment link not found.")
        from app.routers.admin import NEEDS_ASSESSMENT_FORM_KEY, _build_default_needs_assessment_form_config, _normalize_form_config

        form_doc = db.needs_assessment_forms.find_one({"key": NEEDS_ASSESSMENT_FORM_KEY, "is_active": True})
        invitation = _serialize_public_invitation(doc)
        invitation["form"] = _normalize_form_config(form_doc) if form_doc else _build_default_needs_assessment_form_config()
        return invitation
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")


@public_router.post("/needs-assessments/{token}")
def submit_public_needs_assessment(token: str, body: dict[str, Any]):
    """Submit a tokenized student needs assessment form."""
    try:
        db = get_db()
        doc = db.enrollments.find_one({"needs_assessment_invitation.token": token, "flagged_for_mentoring": True})
        if not doc:
            raise HTTPException(status_code=404, detail="Needs assessment link not found.")

        invitation = doc.get("needs_assessment_invitation") or {}
        if doc.get("needs_assessment") and invitation.get("status") == "completed":
            raise HTTPException(status_code=409, detail="This needs assessment form has already been submitted.")

        from app.routers.admin import NEEDS_ASSESSMENT_FORM_KEY, _build_default_needs_assessment_form_config, _normalize_form_config

        form_doc = db.needs_assessment_forms.find_one({"key": NEEDS_ASSESSMENT_FORM_KEY, "is_active": True})
        active_form = _normalize_form_config(form_doc) if form_doc else _build_default_needs_assessment_form_config()
        payload = _normalize_needs_assessment_payload(body or {})
        missing_required_fields = _collect_missing_required_needs_assessment_fields(payload, active_form)
        if missing_required_fields:
            if len(missing_required_fields) == 1:
                detail = f"Please complete the required field: {missing_required_fields[0]}."
            else:
                detail = "Please complete the required fields: " + ", ".join(missing_required_fields) + "."
            raise HTTPException(
                status_code=400,
                detail={
                    "message": detail,
                    "missing_required_fields": missing_required_fields,
                },
            )
        payload["_submitted_via"] = "public_form"
        now = datetime.now(timezone.utc)
        updated_invitation = {
            **invitation,
            "status": "completed",
            "submitted_at": now,
            "last_error": None,
        }
        db.enrollments.update_one(
            {"_id": doc["_id"]},
            {
                "$set": {
                    "needs_assessment": payload,
                    "needs_assessment_uploaded_at": now,
                    "needs_assessment_invitation": updated_invitation,
                    "academic_challenge_score": payload.get("academic_challenge_score"),
                    "external_factor_score": payload.get("external_factor_score"),
                }
            },
        )
        return {
            "message": "Needs assessment submitted successfully.",
            "submitted_at": now.isoformat(),
        }
    except HTTPException:
        raise
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable.")
